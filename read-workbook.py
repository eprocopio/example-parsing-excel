### this is an example of interacting with an Excel workbook using openpyxl

from openpyxl import load_workbook
wb = load_workbook(filename = 'test-workbook.xlsx')
ws = wb.worksheets[0]

# EXAMPLE - READING SINGLE VALUE
ac = ws['A1']
acValue = ac.value
print(acValue)

# EXAMPLE - READING 1-DIMENSIONAL ARRAY
ar = ws['A1':'A3']
arItem = ar[0]
arItemOfItem=arItem[0]
arItemOfItem.value="eric"

# EXAMPLE - READING 2-DIMENSIONAL ARRAY
ar = ws['A1':'C3']
arItem = ar[0][0].value
#arItemOfItem=arItem[0]


mine = 1
